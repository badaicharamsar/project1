from __future__ import annotations

import csv
import difflib
import json
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..schemas import DatasetSchema, LoadDataRequest, TextDataSource

NODE_REQUIRED_FIELDS = {"id", "name", "latitude", "longitude"}
EDGE_REQUIRED_FIELDS = {"source", "target", "time"}
EDGE_OPTIONAL_IMPORT_FIELDS = {"distance", "cost", "road_condition_score"}
EDGE_DEFAULTS = {"cost": 0.0, "road_condition_score": 3.0}
MAX_HEADER_SCAN_ROWS = 10

PREFERRED_NODE_SHEET_NAMES = {
    "kordinat simpul",
    "koordinat simpul",
    "kordinat_simpul",
    "koordinat_simpul",
    "simpul",
}
PREFERRED_EDGE_SHEET_NAMES = {
    "datafix",
    "data fix",
    "data_fix",
    "jalur",
    "data jalur",
    "data_jalur",
}

NODE_HEADER_ALIASES = {
    "id": {"id", "kode simpul", "kode_simpul", "node id", "kode node"},
    "name": {"name", "nama", "nama persimpangan", "nama simpul"},
    "latitude": {"latitude", "lat"},
    "longitude": {"longitude", "long", "lng", "lon"},
    "active": {"active", "aktif"},
}

EDGE_HEADER_ALIASES = {
    "source": {"source", "simpul awal", "node asal", "asal", "from"},
    "target": {"target", "simpul akhir", "node tujuan", "tujuan", "to"},
    "distance": {"distance", "jarak", "jarak km", "jarak (km)"},
    "time": {"time", "waktu", "durasi"},
    "cost": {"cost", "biaya"},
    "road_condition_score": {
        "road_condition_score",
        "road condition",
        "road condition score",
        "kondisi jalan",
        "skor kondisi jalan",
    },
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\n\r\t]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _header_matches_alias(normalized_header: str, alias: str) -> bool:
    if normalized_header == alias:
        return True
    if alias and alias in normalized_header:
        return True
    if normalized_header and normalized_header in alias:
        return True

    header_tokens = set(normalized_header.split())
    alias_tokens = set(alias.split())
    if alias_tokens and alias_tokens.issubset(header_tokens):
        return True

    return False


def _slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "field"


def _normalize_sheet_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _coerce_float(value: Any, field_name: str, row_number: int) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Invalid numeric value for '{field_name}' at row {row_number}: {value!r}."
        ) from exc


def _coerce_metadata_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        numeric_value = float(text)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value
    except ValueError:
        return text


def _resolve_header_map(fieldnames: list[str], record_type: Literal["nodes", "edges"]) -> dict[str, str]:
    alias_map = NODE_HEADER_ALIASES if record_type == "nodes" else EDGE_HEADER_ALIASES
    resolved: dict[str, str] = {}
    taken_canonicals: set[str] = set()

    for original_header in fieldnames:
        normalized_header = _normalize_header(original_header)
        for canonical_name, aliases in alias_map.items():
            if canonical_name in taken_canonicals:
                continue
            if any(_header_matches_alias(normalized_header, alias) for alias in aliases):
                resolved[original_header] = canonical_name
                taken_canonicals.add(canonical_name)
                break

    return resolved


def _has_required_headers(fieldnames: list[str], record_type: Literal["nodes", "edges"]) -> bool:
    header_map = _resolve_header_map(fieldnames, record_type)
    required_fields = NODE_REQUIRED_FIELDS if record_type == "nodes" else EDGE_REQUIRED_FIELDS
    return required_fields.issubset(set(header_map.values()))


def _parse_json_records(content: str, record_type: Literal["nodes", "edges"]) -> list[dict[str, Any]]:
    try:
        loaded = json.loads(content)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON for {record_type}: {exc.msg}.") from exc

    if isinstance(loaded, dict):
        if record_type in loaded:
            records = loaded[record_type]
        else:
            raise ValueError(
                f"JSON for {record_type} must be a list or an object containing the key '{record_type}'."
            )
    elif isinstance(loaded, list):
        records = loaded
    else:
        raise ValueError(f"JSON for {record_type} must be a list of records.")

    if not isinstance(records, list):
        raise ValueError(f"JSON for {record_type} must contain a list of records.")

    return [dict(record) for record in records]


def _build_node_record(row: dict[str, Any], row_number: int, header_map: dict[str, str]) -> dict[str, Any]:
    missing = NODE_REQUIRED_FIELDS - set(header_map.values())
    if missing:
        raise ValueError(
            "Node file is missing required columns. "
            f"Accepted concepts are: {sorted(NODE_REQUIRED_FIELDS)}."
        )

    metadata: dict[str, Any] = {}
    for original_header, value in row.items():
        canonical_name = header_map.get(original_header)
        if canonical_name in NODE_REQUIRED_FIELDS | {"active"}:
            continue
        coerced = _coerce_metadata_value(value)
        if coerced is not None:
            metadata[_slugify(original_header)] = coerced

    active_source_header = next(
        (header for header, canonical_name in header_map.items() if canonical_name == "active"),
        None,
    )
    active_value = row.get(active_source_header, True) if active_source_header else True

    return {
        "id": str(row[next(header for header, canonical_name in header_map.items() if canonical_name == "id")]).strip(),
        "name": str(
            row[next(header for header, canonical_name in header_map.items() if canonical_name == "name")]
        ).strip(),
        "latitude": _coerce_float(
            row[next(header for header, canonical_name in header_map.items() if canonical_name == "latitude")],
            "latitude",
            row_number,
        ),
        "longitude": _coerce_float(
            row[next(header for header, canonical_name in header_map.items() if canonical_name == "longitude")],
            "longitude",
            row_number,
        ),
        "metadata": metadata,
        "active": str(active_value).strip().lower() not in {"false", "0", "no", "tidak"},
    }


def _build_edge_record(row: dict[str, Any], row_number: int, header_map: dict[str, str]) -> dict[str, Any]:
    missing = EDGE_REQUIRED_FIELDS - set(header_map.values())
    if missing:
        raise ValueError(
            "Edge file is missing required columns. "
            f"Accepted concepts are: {sorted(EDGE_REQUIRED_FIELDS)}."
        )

    metadata: dict[str, Any] = {"__column_labels__": {}}
    for original_header, value in row.items():
        canonical_name = header_map.get(original_header)
        if canonical_name in EDGE_REQUIRED_FIELDS | set(EDGE_DEFAULTS.keys()) | {"distance"}:
            if canonical_name:
                metadata["__column_labels__"][canonical_name] = original_header
            continue
        coerced = _coerce_metadata_value(value)
        if coerced is not None:
            slug_key = _slugify(original_header)
            metadata[slug_key] = coerced
            metadata["__column_labels__"][slug_key] = original_header

    def get_value(canonical_name: str, default: Any = None) -> Any:
        source_header = next(
            (header for header, mapped_name in header_map.items() if mapped_name == canonical_name),
            None,
        )
        return row.get(source_header, default) if source_header else default

    distance_value = get_value("distance", None)
    distance_missing = distance_value in (None, "")
    if distance_missing:
        metadata["_missing_distance"] = True
    source_value = str(get_value("source")).strip()
    target_value = str(get_value("target")).strip()
    edge_code = metadata.get("kode_sisi")
    metadata["__edge_id"] = str(edge_code).strip() if edge_code not in (None, "") else f"{source_value}->{target_value}#{row_number}"
    metadata["__edge_label"] = (
        str(edge_code).strip()
        if edge_code not in (None, "")
        else f"{source_value} -> {target_value}"
    )

    return {
        "source": source_value,
        "target": target_value,
        "distance": _coerce_float(0.0 if distance_missing else distance_value, "distance", row_number),
        "time": _coerce_float(get_value("time"), "time", row_number),
        "cost": _coerce_float(get_value("cost", EDGE_DEFAULTS["cost"]), "cost", row_number),
        "road_condition_score": _coerce_float(
            get_value("road_condition_score", EDGE_DEFAULTS["road_condition_score"]),
            "road_condition_score",
            row_number,
        ),
        "metadata": metadata,
    }


def _parse_tabular_records(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    record_type: Literal["nodes", "edges"],
) -> list[dict[str, Any]]:
    header_map = _resolve_header_map(fieldnames, record_type)
    parsed_records: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows, start=2):
        if all(value in (None, "") for value in row.values()):
            continue
        if record_type == "nodes":
            parsed_records.append(_build_node_record(row, row_number, header_map))
        else:
            parsed_records.append(_build_edge_record(row, row_number, header_map))

    return parsed_records


def _parse_nodes_csv(content: str) -> list[dict[str, Any]]:
    reader = csv.DictReader(StringIO(content))
    return _parse_tabular_records(list(reader), list(reader.fieldnames or []), "nodes")


def _parse_edges_csv(content: str) -> list[dict[str, Any]]:
    reader = csv.DictReader(StringIO(content))
    return _parse_tabular_records(list(reader), list(reader.fieldnames or []), "edges")


def _parse_excel_records(content: bytes, record_type: Literal["nodes", "edges"]) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    worksheet_info = _find_worksheet(workbook, record_type)
    return _parse_tabular_records(worksheet_info["rows"], worksheet_info["fieldnames"], record_type)


def _worksheet_to_rows(
    worksheet: Worksheet,
    header_row_index: int = 0,
) -> tuple[list[str], list[dict[str, Any]]]:
    sheet_rows = list(worksheet.iter_rows(values_only=True))
    if not sheet_rows:
        return [], []

    if header_row_index >= len(sheet_rows):
        return [], []

    raw_headers = [cell for cell in sheet_rows[header_row_index]]
    fieldnames = [
        str(cell).strip() if cell is not None else f"column_{index + 1}"
        for index, cell in enumerate(raw_headers)
    ]

    records: list[dict[str, Any]] = []
    for raw_row in sheet_rows[header_row_index + 1 :]:
        row_dict: dict[str, Any] = {}
        for index, header in enumerate(fieldnames):
            row_dict[header] = raw_row[index] if index < len(raw_row) else None
        records.append(row_dict)
    return fieldnames, records


def _detect_header_rows(
    worksheet: Worksheet,
    record_type: Literal["nodes", "edges"],
    max_scan_rows: int = MAX_HEADER_SCAN_ROWS,
) -> dict[str, Any] | None:
    sheet_rows = list(worksheet.iter_rows(values_only=True))
    if not sheet_rows:
        return None

    max_index = min(len(sheet_rows), max_scan_rows)
    for header_row_index in range(max_index):
        fieldnames, records = _worksheet_to_rows(worksheet, header_row_index)
        if fieldnames and _has_required_headers(fieldnames, record_type):
            return {
                "worksheet": worksheet,
                "fieldnames": fieldnames,
                "rows": records,
                "header_row_index": header_row_index,
            }
    return None


def _describe_header_candidates(workbook: Workbook, record_type: Literal["nodes", "edges"]) -> str:
    alias_map = NODE_HEADER_ALIASES if record_type == "nodes" else EDGE_HEADER_ALIASES
    known_aliases = sorted({alias for aliases in alias_map.values() for alias in aliases})
    descriptions: list[str] = []

    for worksheet in workbook.worksheets:
        sheet_rows = list(worksheet.iter_rows(values_only=True))
        preview_lines: list[str] = []
        for row_index in range(min(len(sheet_rows), 3)):
            cells = [
                str(cell).strip()
                for cell in sheet_rows[row_index]
                if cell not in (None, "")
            ]
            if not cells:
                continue
            normalized_cells = [_normalize_header(cell) for cell in cells]
            fuzzy_matches = [
                difflib.get_close_matches(cell, known_aliases, n=1, cutoff=0.75)[0]
                for cell in normalized_cells
                if difflib.get_close_matches(cell, known_aliases, n=1, cutoff=0.75)
            ]
            preview_lines.append(
                f"row {row_index + 1}: headers={cells[:6]} matched_aliases={fuzzy_matches[:6]}"
            )
        if preview_lines:
            descriptions.append(f"sheet '{worksheet.title}' -> " + " | ".join(preview_lines))

    return " ; ".join(descriptions)


def _find_worksheet(
    workbook: Workbook,
    record_type: Literal["nodes", "edges"],
) -> dict[str, Any]:
    preferred_names = (
        PREFERRED_NODE_SHEET_NAMES if record_type == "nodes" else PREFERRED_EDGE_SHEET_NAMES
    )
    preferred_matches: list[dict[str, Any]] = []
    fallback_matches: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        worksheet_info = _detect_header_rows(worksheet, record_type)
        if worksheet_info is None:
            continue

        normalized_name = _normalize_sheet_name(worksheet.title)
        if normalized_name in preferred_names:
            preferred_matches.append(worksheet_info)
        else:
            fallback_matches.append(worksheet_info)

    if preferred_matches:
        return preferred_matches[0]
    if fallback_matches:
        return fallback_matches[0]

    expected_headers = (
        sorted(NODE_REQUIRED_FIELDS)
        if record_type == "nodes"
        else sorted(EDGE_REQUIRED_FIELDS | {"distance"})
    )
    preferred_name_hint = (
        "sheet seperti 'kordinat simpul'"
        if record_type == "nodes"
        else "sheet seperti 'datafix'"
    )
    raise ValueError(
        f"No worksheet matched the expected {record_type} format inside the uploaded workbook. "
        f"Pastikan ada {preferred_name_hint} dan header ditemukan dalam {MAX_HEADER_SCAN_ROWS} baris pertama. "
        f"Konsep kolom yang dicari: {expected_headers}. "
        f"Preview yang terbaca sistem: {_describe_header_candidates(workbook, record_type)}"
    )


def parse_text_source(source: TextDataSource, record_type: Literal["nodes", "edges"]) -> list[dict[str, Any]]:
    if source.format == "json":
        return _parse_json_records(source.content, record_type)
    if source.format == "csv":
        if record_type == "nodes":
            return _parse_nodes_csv(source.content)
        return _parse_edges_csv(source.content)
    raise ValueError(f"Unsupported source format: {source.format}.")


def resolve_dataset(request: LoadDataRequest) -> DatasetSchema:
    """Resolve uploaded text or a direct JSON payload into a validated dataset."""

    if request.dataset is not None:
        return request.dataset

    if request.nodes_source is None or request.edges_source is None:
        raise ValueError(
            "Provide either a complete dataset payload or both nodes_source and edges_source."
        )

    nodes = parse_text_source(request.nodes_source, "nodes")
    edges = parse_text_source(request.edges_source, "edges")
    return DatasetSchema(nodes=nodes, edges=edges)


def resolve_excel_dataset(nodes_content: bytes, edges_content: bytes) -> DatasetSchema:
    nodes = _parse_excel_records(nodes_content, "nodes")
    edges = _parse_excel_records(edges_content, "edges")
    return DatasetSchema(nodes=nodes, edges=edges)


def resolve_single_workbook_dataset(workbook_content: bytes) -> DatasetSchema:
    workbook = load_workbook(filename=BytesIO(workbook_content), data_only=True)
    edge_sheet = _find_worksheet(workbook, "edges")
    node_sheet = _find_worksheet(workbook, "nodes")

    nodes = _parse_tabular_records(node_sheet["rows"], node_sheet["fieldnames"], "nodes")
    edges = _parse_tabular_records(edge_sheet["rows"], edge_sheet["fieldnames"], "edges")
    return DatasetSchema(nodes=nodes, edges=edges)


def load_demo_dataset(demo_dir: Path) -> DatasetSchema:
    nodes_path = demo_dir / "demo_nodes.json"
    edges_path = demo_dir / "demo_edges.json"

    if not nodes_path.exists() or not edges_path.exists():
        raise FileNotFoundError("Demo data files are missing.")

    nodes = _parse_json_records(nodes_path.read_text(encoding="utf-8"), "nodes")
    edges = _parse_json_records(edges_path.read_text(encoding="utf-8"), "edges")
    return DatasetSchema(nodes=nodes, edges=edges)
